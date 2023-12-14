from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import random

# 엑셀 파일 생성
workbook = Workbook()
sheet = workbook.active
sheet.title = "Product Data"

# 열 제목 추가
headers = ['제품ID', '제품명', '수량', '가격']
for col_idx, header in enumerate(headers, start=1):
    sheet.cell(row=1, column=col_idx).value = header

# 제품 데이터 생성 및 추가
for row_idx in range(2, 102):  # 100개의 데이터 생성
    product_id = f'P{row_idx - 1}'
    product_name = f'제품{row_idx - 1}'
    quantity = random.randint(1, 50)
    price = round(random.uniform(10, 1000), 2)

    product_data = [product_id, product_name, quantity, price]

    for col_idx, data in enumerate(product_data, start=1):
        sheet.cell(row=row_idx, column=col_idx).value = data

# 파일 저장
file_path = 'C:\\work\\product.xlsx'
workbook.save(file_path)
